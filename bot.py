# ============================================================
# SECTION 1: IMPORTS
# ============================================================
import random
import json
import os
import io
from urllib.parse import quote
from openpyxl import load_workbook
import qrcode

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    ConversationHandler, ContextTypes, ChatMemberHandler,
    CallbackQueryHandler, filters
)

from config import BOT_TOKEN, ADMIN_IDS, UPI_ID, UPI_NAME


# ============================================================
# SECTION 2: CONSTANTS
# ============================================================
NAME, AGE, CONTACT, EMAIL, DESCRIPTION = range(5)

GROUPS_FILE = "known_groups.json"
LIBRARY_FILE = "library.xlsx"
PURCHASES_FILE = "purchases.json"
PENDING_FILE = "pending_payments.json"


# ============================================================
# SECTION 3: DATA STORAGE HELPERS
# ============================================================
def _atomic_write_json(path, data):
    """
    Writes JSON safely: saves to a temporary file first, then swaps it into
    place with os.replace(). os.replace() is atomic on both Windows and
    Linux — the target file is either fully the old version or fully the
    new version, never a half-written/corrupted mix, even if the process
    is killed, the PC loses power, or the bot crashes mid-save.
    """
    tmp_path = f"{path}.tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)
    os.replace(tmp_path, path)

def load_groups():
    if os.path.exists(GROUPS_FILE):
        with open(GROUPS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_groups(groups):
    _atomic_write_json(GROUPS_FILE, groups)

def load_library():
    """Reads Keyword, Link, Price from library.xlsx. Price defaults to 0 (free) if blank."""
    wb = load_workbook(LIBRARY_FILE)
    ws = wb.active
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        keyword, link = row[0], row[1]
        price = row[2] if len(row) > 2 and row[2] else 0
        entries.append({"keyword": keyword, "link": link, "price": price})
    return entries

def load_purchases():
    if os.path.exists(PURCHASES_FILE):
        with open(PURCHASES_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_purchases(purchases):
    _atomic_write_json(PURCHASES_FILE, purchases)

def has_purchased(user_id, keyword):
    purchases = load_purchases()
    return keyword in purchases.get(str(user_id), [])

def grant_purchase(user_id, keyword):
    purchases = load_purchases()
    user_key = str(user_id)
    purchases.setdefault(user_key, [])
    if keyword not in purchases[user_key]:
        purchases[user_key].append(keyword)
    save_purchases(purchases)

def load_pending():
    """
    Loads pending_payments.json. The request-ID counter lives inside this
    same file (under the reserved key "_next_id") rather than in a
    separate file — that way, allocating a new ID and saving the new
    payment request happen in a single load-modify-save cycle on ONE
    file, instead of two files that could fall out of sync with each
    other if something goes wrong between writing one and writing the
    other.
    """
    if os.path.exists(PENDING_FILE):
        with open(PENDING_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
    else:
        data = {}
    data.setdefault("_next_id", 1)
    return data

def save_pending(pending):
    _atomic_write_json(PENDING_FILE, pending)

def allocate_request_id(pending):
    """
    Reserves and returns the next sequential, human-readable request ID
    (001, 002, 003, ...). Mutates the given `pending` dict in place —
    caller must still call save_pending(pending) afterwards to persist it,
    alongside the new request entry, in one write.
    """
    request_id = f"{pending['_next_id']:03d}"
    pending["_next_id"] += 1
    return request_id


# ============================================================
# SECTION 4: BASIC COMMANDS (start, menu)
# ============================================================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Hey! I'm alive and working.")

async def show_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    buttons = [
        [InlineKeyboardButton("🎫 Raise a Ticket", callback_data="menu:ticket")],
        [
            InlineKeyboardButton("👥 My Groups", callback_data="menu:groups"),
            InlineKeyboardButton("📚 Library", callback_data="menu:library"),
        ],
        [InlineKeyboardButton("💳 My Purchases", callback_data="menu:purchases")],
    ]
    await update.message.reply_text(
        "Here's what I can do:", reply_markup=InlineKeyboardMarkup(buttons)
    )


# ============================================================
# SECTION 5: TICKET SYSTEM
# ============================================================
async def ticket_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['ticket'] = {}
    text = "Let's get your ticket started.\n\nWhat's your Name/Username?"
    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.message.reply_text(text)
    else:
        await update.message.reply_text(text)
    return NAME

async def got_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['ticket']['name'] = update.message.text
    await update.message.reply_text("Got it. What's your Age?")
    return AGE

async def got_age(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['ticket']['age'] = update.message.text
    await update.message.reply_text(
        "Contact number? (Optional — type 'skip' if you'd rather not.)"
    )
    return CONTACT

async def got_contact(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    context.user_data['ticket']['contact'] = "" if text.lower() == "skip" else text
    await update.message.reply_text(
        "Email? (Optional — type 'skip' if you'd rather not.)"
    )
    return EMAIL

async def got_email(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    context.user_data['ticket']['email'] = "" if text.lower() == "skip" else text
    await update.message.reply_text("Last thing — describe your request.")
    return DESCRIPTION

async def got_description(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['ticket']['description'] = update.message.text
    ticket = context.user_data['ticket']
    user = update.effective_user

    summary = (
        f"🎫 New Ticket\n\n"
        f"From: {user.full_name} (@{user.username}, ID: {user.id})\n"
        f"Name/Username given: {ticket['name']}\n"
        f"Age: {ticket['age']}\n"
        f"Contact: {ticket['contact'] or 'Not provided'}\n"
        f"Email: {ticket['email'] or 'Not provided'}\n"
        f"Description: {ticket['description']}"
    )

    for admin_id in ADMIN_IDS:
        await context.bot.send_message(chat_id=admin_id, text=summary)

    await update.message.reply_text(
        "Your ticket's been submitted. An admin will reach out when available. Thanks for your patience!"
    )
    context.user_data.clear()
    return ConversationHandler.END

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text("Ticket cancelled.")
    return ConversationHandler.END


# ============================================================
# SECTION 6: GROUP MANAGEMENT
# ============================================================
async def track_group(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat = update.effective_chat
    new_status = update.my_chat_member.new_chat_member.status
    groups = load_groups()

    if new_status in ("member", "administrator"):
        if str(chat.id) not in groups:
            try:
                invite_link = await context.bot.export_chat_invite_link(chat.id)
            except Exception:
                invite_link = None
            groups[str(chat.id)] = {"title": chat.title, "invite_link": invite_link}
            save_groups(groups)
            print(f"Now tracking group: {chat.title}")
    elif new_status in ("left", "kicked"):
        groups.pop(str(chat.id), None)
        save_groups(groups)
        print(f"No longer in group: {chat.title}")

async def build_groups_response(user_id, context):
    groups = load_groups()
    if not groups:
        return "I'm not managing any groups yet.", None

    buttons = []
    for chat_id, info in groups.items():
        try:
            member = await context.bot.get_chat_member(chat_id, user_id)
            is_in = member.status in ("member", "administrator", "creator")
        except Exception:
            is_in = False

        if is_in:
            buttons.append([InlineKeyboardButton(
                f"🚪 Leave {info['title']}", callback_data=f"leave:{chat_id}"
            )])
        elif info.get("invite_link"):
            buttons.append([InlineKeyboardButton(
                f"➕ Join {info['title']}", url=info["invite_link"]
            )])

    if not buttons:
        return "No groups to show right now.", None
    return "Here are the groups I manage:", InlineKeyboardMarkup(buttons)

async def my_groups(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text, markup = await build_groups_response(update.effective_user.id, context)
    await update.message.reply_text(text, reply_markup=markup)

async def menu_groups_button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    text, markup = await build_groups_response(query.from_user.id, context)
    await query.message.reply_text(text, reply_markup=markup)

async def leave_group_button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    chat_id = query.data.split(":")[1]
    user_id = query.from_user.id

    try:
        await context.bot.ban_chat_member(chat_id, user_id)
        await context.bot.unban_chat_member(chat_id, user_id, only_if_banned=True)
        await query.edit_message_text("You've been removed from that group.")
    except Exception as e:
        await query.edit_message_text(f"Couldn't remove you: {e}")


# ============================================================
# SECTION 7: LIBRARY SEARCH (now price-aware)
# ============================================================
def build_entry_button(entry, user_id):
    """
    Returns one button for a library entry.
    - Free, or already purchased -> opens the link directly.
    - Paid and not yet purchased -> triggers the payment flow instead.
    """
    keyword, link, price = entry["keyword"], entry["link"], entry["price"]

    if price and price > 0 and not has_purchased(user_id, keyword):
        label = f"🔒 {keyword} (₹{price})"
        return InlineKeyboardButton(label, callback_data=f"pay:{keyword}")
    else:
        label = f"🔓 {keyword}" if price and price > 0 else keyword
        return InlineKeyboardButton(label, url=link)

async def library_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    buttons = [
        [InlineKeyboardButton("🎲 Random", callback_data="lib:random")],
        [InlineKeyboardButton("🔤 Word Based", callback_data="lib:word")],
    ]
    await update.message.reply_text(
        "How would you like to search?", reply_markup=InlineKeyboardMarkup(buttons)
    )

async def menu_library_button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    buttons = [
        [InlineKeyboardButton("🎲 Random", callback_data="lib:random")],
        [InlineKeyboardButton("🔤 Word Based", callback_data="lib:word")],
    ]
    await query.message.reply_text(
        "How would you like to search?", reply_markup=InlineKeyboardMarkup(buttons)
    )

async def library_button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    if query.data == "lib:random":
        entries = load_library()
        if not entries:
            await query.edit_message_text("Library is empty right now.")
            return
        pick = random.choice(entries)
        button = build_entry_button(pick, user_id)
        await query.edit_message_text(
            f"🎲 {pick['keyword']}",
            reply_markup=InlineKeyboardMarkup([[button]])
        )

    elif query.data == "lib:word":
        await query.edit_message_text("Type the letter you want to search by (e.g. 'c').")
        context.user_data['awaiting_letter'] = True

async def handle_letter_search(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.user_data.get('awaiting_letter'):
        return

    context.user_data['awaiting_letter'] = False
    letter = update.message.text.strip().lower()
    user_id = update.effective_user.id

    entries = load_library()
    matches = [e for e in entries if e['keyword'].lower().startswith(letter)]

    if not matches:
        await update.message.reply_text(f"Nothing found starting with '{letter}'.")
        return

    buttons = [[build_entry_button(e, user_id)] for e in matches]
    await update.message.reply_text(
        f"Results for '{letter}':", reply_markup=InlineKeyboardMarkup(buttons)
    )

def build_purchases_response(user_id):
    """
    Builds the 'My Purchases' view: only the paid items THIS user has
    actually unlocked, matched against the current library so removed/
    renamed items don't show a stale or broken link.
    """
    purchases = load_purchases()
    owned_keywords = purchases.get(str(user_id), [])

    if not owned_keywords:
        return "You haven't purchased anything yet.", None

    entries = load_library()
    owned_entries = [e for e in entries if e["keyword"] in owned_keywords]

    if not owned_entries:
        return "You haven't purchased anything yet.", None

    # Reuses build_entry_button — since these are already purchased,
    # has_purchased() will be true for each one, so it renders as an
    # unlocked 🔓 direct-link button rather than a locked/pay one.
    buttons = [[build_entry_button(e, user_id)] for e in owned_entries]
    return "💳 Your purchased items:", InlineKeyboardMarkup(buttons)

async def my_purchases(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text, markup = build_purchases_response(update.effective_user.id)
    await update.message.reply_text(text, reply_markup=markup)

async def menu_purchases_button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    text, markup = build_purchases_response(query.from_user.id)
    await query.message.reply_text(text, reply_markup=markup)


# ============================================================
# SECTION 8: PAYMENTS (manual UPI + admin confirmation)
# ============================================================
async def pay_button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """User tapped a locked (🔒) library entry."""
    query = update.callback_query
    await query.answer()

    keyword = query.data.split(":", 1)[1]
    entries = load_library()
    entry = next((e for e in entries if e["keyword"] == keyword), None)

    if not entry:
        await query.message.reply_text("Sorry, that item isn't available anymore.")
        return

    user_id = query.from_user.id

    # Guard against duplicate payment requests: if this user already owns
    # the item (e.g. they tapped an old 🔒 button still sitting in their
    # chat history from before it was confirmed), just hand them the link
    # instead of spinning up a brand-new payment — avoids anyone paying
    # twice for the same item and running into bank-side duplicate-charge
    # hassles.
    if has_purchased(user_id, entry["keyword"]):
        await query.message.reply_text(
            f"✅ You've already purchased {entry['keyword']} — here's your link:\n\n{entry['link']}"
        )
        return

    pending = load_pending()
    request_id = allocate_request_id(pending)
    pending[request_id] = {
        "user_id": user_id,
        "username": query.from_user.username or query.from_user.full_name,
        "keyword": entry["keyword"],
        "price": entry["price"],
        "status": "awaiting_proof",
    }
    save_pending(pending)

    # Remember every request this user currently has open, so that if they
    # pay for two different items back-to-back, a screenshot for one
    # doesn't get mistaken for the other.
    context.user_data.setdefault('awaiting_proof_ids', [])
    context.user_data['awaiting_proof_ids'].append(request_id)

    note = quote(f"Library-{entry['keyword']}")
    upi_link = f"upi://pay?pa={UPI_ID}&pn={quote(UPI_NAME)}&am={entry['price']}&cu=INR&tn={note}"

    text = (
        f"🔒 {entry['keyword']} — ₹{entry['price']}\n"
        f"Request ID: {request_id}\n\n"
        f"This item is locked — it's a paid entry.\n\n"
        f"To unlock it, pay ₹{entry['price']} via UPI:\n"
        f"• Scan the QR code below with any UPI app (GPay, PhonePe, Paytm, etc.), OR\n"
        f"• Pay manually to this UPI ID: {UPI_ID}\n\n"
        f"Once you've paid, send a SCREENSHOT of the payment right here in this chat as proof. "
        f"An admin will review it and unlock the item.\n\n"
        f"If you're paying for more than one item at a time, add the Request ID "
        f"({request_id}) as the photo's caption so we know which one it's for."
    )

    # Telegram inline URL buttons only accept http(s):// or tg:// links —
    # a upi:// deep link is rejected outright (that's the BadRequest you hit).
    # A QR code sidesteps this entirely: any UPI app can scan it directly,
    # no clickable link required.
    qr_image = qrcode.make(upi_link)
    qr_buffer = io.BytesIO()
    qr_image.save(qr_buffer, format="PNG")
    qr_buffer.seek(0)

    await query.message.reply_photo(photo=qr_buffer, caption=text)

async def handle_payment_proof(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """User sends a photo while one or more payments are awaiting proof. Forwards it to admins for review."""
    open_ids = context.user_data.get('awaiting_proof_ids', [])
    pending = load_pending()

    # Only keep IDs that are still genuinely waiting on proof (filters out
    # stale/already-handled ones left over from earlier in the session).
    active_ids = [rid for rid in open_ids if pending.get(rid, {}).get("status") == "awaiting_proof"]
    context.user_data['awaiting_proof_ids'] = active_ids

    if not active_ids:
        # Not currently in the middle of a payment — ignore the photo entirely.
        return

    if len(active_ids) == 1:
        request_id = active_ids[0]
    else:
        # More than one payment open at once — need the caption to tell them apart.
        caption = (update.message.caption or "").strip()
        if caption in active_ids:
            request_id = caption
        else:
            options = "\n".join(f"• {rid} — {pending[rid]['keyword']}" for rid in active_ids)
            await update.message.reply_text(
                "You've got more than one payment waiting on proof:\n\n"
                f"{options}\n\n"
                "Please resend the screenshot with the matching Request ID typed as the photo's caption."
            )
            return

    request = pending[request_id]

    # Highest-resolution version of the photo the user sent.
    photo_file_id = update.message.photo[-1].file_id

    request["status"] = "pending_admin_review"
    request["photo_file_id"] = photo_file_id
    pending[request_id] = request
    save_pending(pending)

    context.user_data['awaiting_proof_ids'].remove(request_id)
    await update.message.reply_text(
        f"Got your screenshot for Request {request_id}! Waiting for admin confirmation..."
    )

    admin_caption = (
        f"💰 Payment claim\n\n"
        f"User: @{request['username']} (ID: {request['user_id']})\n"
        f"Item: {request['keyword']}\n"
        f"Amount: ₹{request['price']}\n"
        f"Request ID: {request_id}"
    )
    admin_buttons = [[
        InlineKeyboardButton("✅ Confirm", callback_data=f"confirm:{request_id}"),
        InlineKeyboardButton("❌ Reject", callback_data=f"reject:{request_id}"),
    ]]
    for admin_id in ADMIN_IDS:
        await context.bot.send_photo(
            chat_id=admin_id, photo=photo_file_id, caption=admin_caption,
            reply_markup=InlineKeyboardMarkup(admin_buttons)
        )

async def confirm_payment(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin taps Confirm — unlocks the item and messages the buyer."""
    query = update.callback_query
    await query.answer()

    if query.from_user.id not in ADMIN_IDS:
        await query.answer("Admins only.", show_alert=True)
        return

    request_id = query.data.split(":", 1)[1]
    pending = load_pending()
    request = pending.get(request_id)

    if not request or request.get("status") != "pending_admin_review":
        # Already handled (e.g. another admin already confirmed/rejected it) or never existed.
        await query.edit_message_caption("This request has already been handled or no longer exists.")
        return

    # Pop it immediately so a second, near-simultaneous admin tap can't process it twice.
    pending.pop(request_id, None)
    save_pending(pending)

    grant_purchase(request["user_id"], request["keyword"])

    entries = load_library()
    entry = next((e for e in entries if e["keyword"] == request["keyword"]), None)
    link = entry["link"] if entry else "(link unavailable — item may have been removed)"

    await context.bot.send_message(
        chat_id=request["user_id"],
        text=f"✅ Payment confirmed! Here's your link:\n\n{request['keyword']}\n{link}"
    )
    await query.edit_message_caption(f"Confirmed. {request['keyword']} unlocked for @{request['username']}.")

async def reject_payment(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin taps Reject — notifies the buyer, no unlock happens."""
    query = update.callback_query
    await query.answer()

    if query.from_user.id not in ADMIN_IDS:
        await query.answer("Admins only.", show_alert=True)
        return

    request_id = query.data.split(":", 1)[1]
    pending = load_pending()
    request = pending.get(request_id)

    if not request or request.get("status") != "pending_admin_review":
        await query.edit_message_caption("This request has already been handled or no longer exists.")
        return

    pending.pop(request_id, None)
    save_pending(pending)

    await context.bot.send_message(
        chat_id=request["user_id"],
        text=f"Your payment for {request['keyword']} couldn't be confirmed. "
             f"Please contact an admin if you believe this is a mistake."
    )
    await query.edit_message_caption(f"Rejected payment claim for @{request['username']} ({request['keyword']}).")


# ============================================================
# SECTION 9: MAIN / HANDLER REGISTRATION
# ============================================================
def main():
    app = Application.builder().token(BOT_TOKEN).build()

    # --- Ticket conversation (registered first, see Section 7 note) ---
    ticket_conv = ConversationHandler(
        entry_points=[
            CommandHandler("ticket", ticket_start),
            CallbackQueryHandler(ticket_start, pattern="^menu:ticket$"),
        ],
        states={
            NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_name)],
            AGE: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_age)],
            CONTACT: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_contact)],
            EMAIL: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_email)],
            DESCRIPTION: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_description)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False,
    )
    app.add_handler(ticket_conv)

    # --- Basic commands ---
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("menu", show_menu))

    # --- Group management ---
    app.add_handler(ChatMemberHandler(track_group, ChatMemberHandler.MY_CHAT_MEMBER))
    app.add_handler(CommandHandler("mygroups", my_groups))
    app.add_handler(CallbackQueryHandler(menu_groups_button, pattern="^menu:groups$"))
    app.add_handler(CallbackQueryHandler(leave_group_button, pattern="^leave:"))

    # --- Library search ---
    app.add_handler(CommandHandler("library", library_start))
    app.add_handler(CallbackQueryHandler(menu_library_button, pattern="^menu:library$"))
    app.add_handler(CallbackQueryHandler(library_button, pattern="^lib:"))

    # --- My Purchases ---
    app.add_handler(CommandHandler("purchases", my_purchases))
    app.add_handler(CallbackQueryHandler(menu_purchases_button, pattern="^menu:purchases$"))

    # --- Payments ---
    app.add_handler(CallbackQueryHandler(pay_button, pattern="^pay:"))
    app.add_handler(CallbackQueryHandler(confirm_payment, pattern="^confirm:"))
    app.add_handler(CallbackQueryHandler(reject_payment, pattern="^reject:"))
    app.add_handler(MessageHandler(filters.PHOTO, handle_payment_proof))

    # This generic text catcher MUST be added last.
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_letter_search))

    print("Bot is running... press Ctrl+C to stop.")
    app.run_polling()


if __name__ == "__main__":
    main()
